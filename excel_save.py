# %%
import openpyxl


# %%
# [[n1,n2],[n3], ...] 이런 꼴로 묶인 결과를 엑셀 파일로 저장
# saver는 직접 쓸 일 없음. save_as_excel만 사용
def saver(news_set, given_row=1, given_col=1, wb=None, ws=None):
    row, col = given_row, given_col
    if type(news_set[0]) == dict:
        for i in range(len(news_set)):
            news = news_set[i]
            ws.cell(row=row, column=col).value = news['title']
            ws.cell(row=row, column=col+1).value = news['published_at']
            ws.cell(row=row, column=col+2).value = news['hilight']
            row += 1
        return row
    else:
        for i in range(len(news_set)):
            ws.cell(row=row, column=col).value = f"Category {i}"
            row = saver(news_set[i], row+1, col+1, wb, ws)
        return row


def save_as_excel(company_name, news_set, directory):
    wb = openpyxl.load_workbook(directory)
    ws = wb.create_sheet(f"{company_name}")
    saver(news_set, wb=wb, ws=ws)
    wb.save(directory)
    wb.close()
